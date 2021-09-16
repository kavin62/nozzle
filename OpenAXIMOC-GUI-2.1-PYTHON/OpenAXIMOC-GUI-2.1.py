from tkinter import *
import sys
import threading as thrd
import matplotlib
matplotlib.use('Agg')
from tkinter.ttk import Progressbar
from tkinter.ttk import Frame
from tkinter.ttk import Notebook
from time import sleep
from mpmath import *
import sympy as sp
import numpy as np
import datetime
import math
import time
import matplotlib.pyplot as plt
from decimal import *
from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles import Font, Color, colors
import matplotlib.colors
import matplotlib.pylab as pyl
import warnings
from tkinter import *
from matplotlib.backends.backend_tkagg import (FigureCanvasTkAgg, NavigationToolbar2Tk)
from matplotlib.backend_bases import key_press_handler
from matplotlib.figure import Figure
import threading as thrd
import pyglet
import os
import pkg_resources.py2_warn
warnings.filterwarnings("ignore")
plt.ioff()
choice = int(1)
root = Tk()
tabControl = Notebook(root)
calcnum = int(0)

tab1 = Frame(tabControl)
tab3 = Frame(tabControl)
tab2 = Frame(tabControl)


tabControl.add(tab1, text ='Calculator')
tabControl.add(tab3, text='Plot')
tabControl.add(tab2, text ='Help')
tabControl.grid(row=0, column=0, sticky=W)

r_cset = IntVar()

j = 0

root.title("Axisymmetric MOC - Nozzle Contour Designer")
root.geometry("760x930")
root.resizable(width=False, height=False)
root.iconbitmap('icon.ico')
moctitle = Label(tab1, text="Axisymmetric MOC Nozzle Countour Designer", font=("Calibri", 20)).grid(row=1,column=0,columnspan=100, pady=30)
author = Label(root, text="Developed by Kyril Palaveev (June 2020), Sheffield Hallam University, UK",font=("Calibri", 12)).grid(row=500, column=0,columnspan=100, pady=10)


def on_close():
    matplotlib.use('TkAgg')
    from tkinter import messagebox
    close = messagebox.askokcancel("Axisymmetric MOC - Nozzle Contour Designer", "Would you like to close the program?")
    if close:
        os._exit(0)
        root.destroy()
root.protocol("WM_DELETE_WINDOW",  on_close)

def parchoice():
    global choice
    choice = int(1)

    parabolic = Button(tab1, text="Parabolic", padx=5, bg="green3", fg="gray99", state=DISABLED,command=parchoice).grid(row=9, column=3, sticky=W)  # , command=parchoice
    cubic = Button(tab1, text="Cubic", padx=5, bg="gray95", command=cubchoice).grid(row=9, column=3, sticky=E)
    q_factor = Entry(tab1, borderwidth=1, width=20, state=DISABLED)
    q_factor.grid(row=10, column=3, pady=5, columnspan=2)

def cubchoice():
    global choice
    choice = int(2)
    parabolic = Button(tab1, text="Parabolic", padx=5, bg="gray95",command=parchoice).grid(row=9, column=3, sticky=W)  # , command=parchoice
    cubic = Button(tab1, text="Cubic", padx=5, bg="green3", fg="gray99", command=cubchoice, state=DISABLED).grid(row=9, column=3, sticky=E)
    q_factor = Entry(tab1, borderwidth=1, width=20)
    q_factor.grid(row=10, column=3, pady=5, columnspan=2)

def r_cchoice():

    if r_cset.get() == 1:
        global r_c
        r_c.insert(0, 1)
        r_c = Entry(tab1, borderwidth=1, width=20, state=DISABLED)
        r_c.grid(row=12, column=1, pady=5)

    else:
        r_c = Entry(tab1, borderwidth=1, width=20)
        r_c.grid(row=12, column=1, pady=5)

if choice == 1:
    q_factor_value = 0
else:
    q_factor_value = float(q_factor.get())


def AXIMOCdesigner(p_e, F_T, gamma, M_e, T_c, R, r_c, num_of_char_lines, numiter, theta_max_rad, L, pressure_type, q_factor):

    global outbox

    fig = Figure(figsize=(16, 9), dpi=200)



    c_v = float(R / (gamma - 1))
    c_p = float(c_v + R)
    p_c = float(p_e/(1+((gamma-1)/2)*M_e**2)**(-gamma/(gamma-1)))
    rho_c = float(p_c/(R*T_c))
    rho_e = float(rho_c*(1+((gamma-1)/2)*M_e**2)**(-1/(gamma-1)))
    p_t = float(p_c * ((2 / (gamma + 1)) ** (gamma / (gamma - 1))))
    T_e = float(T_c/(p_c/p_e)**((gamma-1)/gamma))
    u_e = M_e * sp.sqrt(gamma * R * T_e)
    A_e = float(F_T / (u_e ** 2 * rho_e))
    M_t = float(1.0001)
    x_t = float(0)
    A_t = float(A_e/sp.sqrt((1/M_e**2)*((2/(gamma+1))*(1+((gamma-1)/2)*M_e**2))**((gamma+1)/(gamma-1))))
    r_e = float(sp.sqrt(A_e/sp.pi))
    r_t = float(sp.sqrt(A_t/sp.pi))

    if r_cset.get() == 1:
        r_c = float(sp.sqrt(3)*r_t)


    dec3 = Decimal(10) ** -3
    dec4 = Decimal(10) ** -4
    dec1 = Decimal(10) ** -2

    warnings.filterwarnings("ignore")

    # _________________________________MOC_________________________________________

    book = openpyxl.load_workbook('export.xlsx')
    sheet = book.active

    v_PMerad = float(sp.sqrt((gamma + 1) / (gamma - 1))) * sp.atan(
        sp.sqrt((gamma - 1) * (M_e ** 2 - 1) / (gamma + 1))) - sp.atan(sp.sqrt(M_e ** 2 - 1))
    v_PMedeg = float(math.degrees(v_PMerad))

    theta_max_deg = float(math.degrees(theta_max_rad))

    first_angle_deg = float(0.001)
    first_angle_rad = float(math.radians(first_angle_deg))

    first_point = float(sp.tan(first_angle_rad) * r_t)

    delta_theta = float((theta_max_rad - first_angle_rad) / (num_of_char_lines - 1))
    delta_theta_deg = float(math.degrees(delta_theta))

    # _____________________________________________________________________________

    z = 0

    x_list = []
    y_list = []
    p_xx_list = []
    M_xx_list = []
    v_PM_xx_rad_list = []
    T_xx_list = []
    V_xx_list = []
    mu_xx_rad_list = []
    theta_xx_rad_list = []
    K_neg_xx_rad_list = []
    K_pos_xx_rad_list = []

    K_neg_xx_deg_list = []
    K_pos_xx_deg_list = []
    theta_xx_deg_list = []
    v_PM_xx_deg_list = []
    mu_xx_deg_list = []

    xerr_list = []
    yerr_list = []
    thetaerr_list = []
    Verr_list = []
    iteration_list = []

    sheet.cell(row=1, column=1).value = "K_neg"
    sheet.cell(row=1, column=2).value = "K_pos"
    sheet.cell(row=1, column=3).value = "Theta"
    sheet.cell(row=1, column=4).value = "v"
    sheet.cell(row=1, column=5).value = "M"
    sheet.cell(row=1, column=6).value = "mu"
    sheet.cell(row=1, column=7).value = "V"
    sheet.cell(row=1, column=8).value = "x"
    sheet.cell(row=1, column=9).value = "y"
    sheet.cell(row=1, column=10).value = "p"
    sheet.cell(row=1, column=11).value = "T"


    if num_of_char_lines > 80:

        outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="yellow")
        outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)

        outbox.insert(END,
                      "CAUTION: For number of characteristic lines above 80, the plot window and plot resizing functions may take longer to respond. It is recommended that the user is patient." + '\n')
        outbox.see("end")

    for x in np.linspace(first_point, L, num=num_of_char_lines, endpoint=True):

        x_list.append(x)

        y = float(0)
        y_list.append(y)

        if pressure_type == 'parabolic':
            p_xx = float(mp.e ** ((L ** (-2)) * (x - L) ** 2 * (sp.ln(p_t) - sp.ln(p_e)) + ln(p_e)))

        if pressure_type == 'cubic':
            p_xx = float(sp.exp(((q_factor * L + 2 * (sp.ln(p_t) - sp.ln(p_e))) / L ** 3) * x ** 3 - ((2 * q_factor * L + 3 * (sp.ln(p_t) - sp.ln(p_e))) / L ** 2) * x ** 2 + q_factor * x + sp.ln(p_t)))

        p_xx_list.append(p_xx)

        M_xx = float(sp.sqrt(((p_c / p_xx) ** ((gamma - 1) / gamma) - 1) * (2 / (gamma - 1))))
        M_xx_list.append(M_xx)


        v_PM_xx_rad = float((math.sqrt((gamma + 1) / (gamma - 1))) * sp.atan(
            math.sqrt((gamma - 1) * (M_xx ** 2 - 1) / (gamma + 1))) - sp.atan(math.sqrt(M_xx ** 2 - 1)))
        v_PM_xx_rad_list.append(v_PM_xx_rad)

        T_xx = T_c * (p_xx / p_c) ** ((gamma - 1) / gamma)
        T_xx_list.append(T_xx)

        V_xx = float(M_xx * sp.sqrt(gamma * R * T_xx))
        V_xx_list.append(V_xx)

        mu_xx_rad = float(sp.asin(1 / M_xx))
        mu_xx_rad_list.append(mu_xx_rad)

        if x == first_point:
            theta_xx_rad = float(first_angle_rad)
        else:
            theta_xx_rad = float(0)

        theta_xx_rad_list.append(theta_xx_rad)

        K_neg_xx_rad = float(theta_xx_rad + v_PM_xx_rad)
        K_neg_xx_rad_list.append(K_neg_xx_rad)

        K_pos_xx_rad = float(theta_xx_rad - v_PM_xx_rad)
        K_pos_xx_rad_list.append(K_pos_xx_rad)

        plt.plot(x_list, y_list, 'k.', markersize=1)

        z += 1

    w = 1
    for g in range(0, z, 1):

        sheet.cell(row=w + 1, column=1).value = math.degrees(K_neg_xx_rad_list[g])
        sheet.cell(row=w + 1, column=2).value = math.degrees(K_pos_xx_rad_list[g])
        sheet.cell(row=w + 1, column=3).value = math.degrees(theta_xx_rad_list[g])
        sheet.cell(row=w + 1, column=4).value = math.degrees(v_PM_xx_rad_list[g])
        sheet.cell(row=w + 1, column=5).value = M_xx_list[g]
        sheet.cell(row=w + 1, column=6).value = math.degrees(mu_xx_rad_list[g])
        sheet.cell(row=w + 1, column=7).value = V_xx_list[g]
        sheet.cell(row=w + 1, column=8).value = x_list[g]
        sheet.cell(row=w + 1, column=9).value = y_list[g]
        sheet.cell(row=w + 1, column=10).value = p_xx_list[g]
        sheet.cell(row=w + 1, column=11).value = T_xx_list[g]

        w = w + z + 1 - g

    for j in range(z - 1, 0, -1):

        f = 0

        progress['value'] = 0
        perbox.delete(0, END)
        tab1.update_idletasks()
        percentage = int(round(float(abs((j/(z)*100)-100))))
        percentagestr = str(percentage)
        strtotal = str(percentagestr+str('%'))
        perbox.insert(0, strtotal)
        progress['value'] = percentage
        tab1.update_idletasks()

        if j == 1:
            perbox.delete(0, END)
            progress['value'] = 100
            perbox.insert(0, 'Done')
            strtotal = str('')
            tab1.update_idletasks()


        remaining.delete(0, j)
        remaining.insert(0, j)

        for i in range(0, j, 1):


            if j == z - 1:

                x = float((y_list[i] - y_list[i + 1] + x_list[i + 1] * sp.tan(
                    theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) - x_list[i] * sp.tan(
                    theta_xx_rad_list[i] + mu_xx_rad_list[i])) / (
                                  sp.tan(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) - sp.tan(
                              theta_xx_rad_list[i] + mu_xx_rad_list[i])))

                y = float(y_list[i] + (x - x_list[i]) * sp.tan(theta_xx_rad_list[i] + mu_xx_rad_list[i]))

                A = float((sp.tan(mu_xx_rad_list[i]) * sp.sin(mu_xx_rad_list[i]) * sp.sin(theta_xx_rad_list[i])) / (
                        y * sp.cos(theta_xx_rad_list[i] + mu_xx_rad_list[i])))
                B = float((sp.tan(mu_xx_rad_list[i + 1]) * sp.sin(mu_xx_rad_list[i + 1]) * sp.sin(
                    theta_xx_rad_list[i + 1])) / (y * sp.cos(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1])))

                V = float(((sp.cot(mu_xx_rad_list[i]) * (1 + (A * (x - x_list[i])))) + (
                        sp.cot(mu_xx_rad_list[i + 1]) * (1 + (B * (x - x_list[i + 1])))) + theta_xx_rad_list[
                               i + 1] - theta_xx_rad_list[i]) / ((sp.cot(mu_xx_rad_list[i]) / V_xx_list[i]) + (
                        sp.cot(mu_xx_rad_list[i + 1]) / V_xx_list[i + 1])))

                theta = float(theta_xx_rad_list[i] + sp.cot(mu_xx_rad_list[i]) * (
                        ((V - V_xx_list[i]) / V_xx_list[i]) - (A * (x - x_list[i]))))

                T = float(T_c - (V ** 2 / (2 * c_p)))

                a = float(sp.sqrt(gamma * R * T))

                M = float(V / a)

                mu = float(sp.asin(1 / M))

                v = float((math.sqrt((gamma + 1) / (gamma - 1))) * sp.atan(
                    math.sqrt((gamma - 1) * (M ** 2 - 1) / (gamma + 1))) - sp.atan(math.sqrt(M ** 2 - 1)))

                K_neg = float(theta + v)
                K_pos = float(theta - v)

                counting = 1
                # --------------------iterations
                for iteration in range(1, numiter + 1, 1):

                    x_2 = float((y_list[i] - y_list[i + 1] + x_list[i + 1] * 0.5 * (
                            sp.tan(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) + sp.tan(theta - mu)) -
                                 x_list[i] * 0.5 * (sp.tan(theta_xx_rad_list[i] + mu_xx_rad_list[i]) + sp.tan(
                                theta + mu))) / (0.5 * (
                            sp.tan(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) - sp.tan(
                        theta_xx_rad_list[i] + mu_xx_rad_list[i]) + sp.tan(theta - mu) - sp.tan(theta + mu))))

                    y_2 = float(y_list[i] + (x - x_list[i]) * 0.5 * (
                            sp.tan(theta_xx_rad_list[i] + mu_xx_rad_list[i]) + sp.tan(theta + mu)))

                    A_2A = float(
                        (sp.tan(mu_xx_rad_list[i]) * sp.sin(mu_xx_rad_list[i]) * sp.sin(theta_xx_rad_list[i])) / (
                                y * sp.cos(theta_xx_rad_list[i] + mu_xx_rad_list[i])))

                    A_2C = float((sp.tan(mu) * sp.sin(mu) * sp.sin(theta)) / (y * sp.cos(theta + mu)))

                    B_2B = float((sp.tan(mu_xx_rad_list[i + 1]) * sp.sin(mu_xx_rad_list[i + 1]) * sp.sin(
                        theta_xx_rad_list[i + 1])) / (y * sp.cos(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1])))

                    B_2C = float((sp.tan(mu) * sp.sin(mu) * sp.sin(theta)) / (y * sp.cos(theta - mu)))

                    A_2 = float((2 / (sp.tan(mu_xx_rad_list[i]) + sp.tan(mu))) * (
                            ((2 * V_xx_list[i]) / (V_xx_list[i] + V)) + ((A_2A + A_2C) / 2) * (x - x_list[i])))

                    B_2 = float((2 / (sp.tan(mu_xx_rad_list[i + 1]) + sp.tan(mu))) * (
                            ((2 * V_xx_list[i + 1]) / (V_xx_list[i + 1] + V)) + ((B_2B + B_2C) / 2) * (
                            x - x_list[i + 1])))

                    V_2 = float((A_2 + B_2 + theta_xx_rad_list[i + 1] - theta_xx_rad_list[i]) / (4 * (
                            (V_xx_list[i] + V) ** (-1) * (sp.tan(mu_xx_rad_list[i]) + sp.tan(mu)) ** (-1) + (
                            V_xx_list[i + 1] + V) ** (-1) * (
                                    sp.tan(mu_xx_rad_list[i + 1]) + sp.tan(mu)) ** (-1))))

                    theta_2 = float(theta_xx_rad_list[i] + (2 / (sp.tan(mu_xx_rad_list[i]) + sp.tan(mu))) * (
                            ((2 * (V_2 - V_xx_list[i])) / (V + V_xx_list[i])) - ((A_2A + A_2C) / 2) * (
                            x - x_list[i])))

                    T_2 = float(T_c - (V_2 ** 2 / (2 * c_p)))

                    a_2 = float(sp.sqrt(gamma * R * T_2))

                    M_2 = float(V_2 / a_2)

                    mu_2 = float(sp.asin(1 / M_2))

                    v_2 = float((math.sqrt((gamma + 1) / (gamma - 1))) * sp.atan(
                        math.sqrt((gamma - 1) * (M_2 ** 2 - 1) / (gamma + 1))) - sp.atan(math.sqrt(M_2 ** 2 - 1)))

                    K_neg_2 = float(theta_2 + v_2)
                    K_pos_2 = float(theta_2 - v_2)

                    xdiff = float(x - x_2)
                    if xdiff == 0:
                        xerr = float(0)
                    else:
                        xerr = float(x_2 / xdiff)

                    ydiff = float(y - y_2)
                    if ydiff == 0:
                        yerr = float(0)
                    else:
                        yerr = float(y_2 / ydiff)

                    thetadiff = float(theta - theta_2)
                    if thetadiff == 0:
                        thetaerr = float(0)
                    else:
                        thetaerr = float(theta_2 / thetadiff)

                    Vdiff = float(V - V_2)
                    if Vdiff == 0:
                        Verr = float(0)
                    else:
                        Verr = float(V_2 / Vdiff)

                    x = x_2
                    y = y_2
                    V = V_2
                    theta = theta_2
                    T = T_2
                    a = a_2
                    M = M_2
                    mu = mu_2
                    v = v_2
                    K_neg = K_neg_2
                    K_pos = K_pos_2

                    counting += 1
            else:

                x = float((y_list[i] - y_list[i + 1] + x_list[i + 1] * sp.tan(
                    theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) - x_list[i] * sp.tan(
                    theta_xx_rad_list[i] + mu_xx_rad_list[i])) / (
                                  sp.tan(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) - sp.tan(
                              theta_xx_rad_list[i] + mu_xx_rad_list[i])))

                y = float(y_list[i] + (x - x_list[i]) * sp.tan(theta_xx_rad_list[i] + mu_xx_rad_list[i]))

                A = float((sp.tan(mu_xx_rad_list[i]) * sp.sin(mu_xx_rad_list[i]) * sp.sin(theta_xx_rad_list[i])) / (
                        y_list[i] * sp.cos(theta_xx_rad_list[i] + mu_xx_rad_list[i])))
                B = float((sp.tan(mu_xx_rad_list[i + 1]) * sp.sin(mu_xx_rad_list[i + 1]) * sp.sin(
                    theta_xx_rad_list[i + 1])) / (
                                  y_list[i + 1] * sp.cos(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1])))

                V = float(((sp.cot(mu_xx_rad_list[i]) * (1 + (A * (x - x_list[i])))) + (
                        sp.cot(mu_xx_rad_list[i + 1]) * (1 + (B * (x - x_list[i + 1])))) + theta_xx_rad_list[
                               i + 1] - theta_xx_rad_list[i]) / ((sp.cot(mu_xx_rad_list[i]) / V_xx_list[i]) + (
                        sp.cot(mu_xx_rad_list[i + 1]) / V_xx_list[i + 1])))

                theta = float(theta_xx_rad_list[i] + sp.cot(mu_xx_rad_list[i]) * (
                        ((V - V_xx_list[i]) / V_xx_list[i]) - (A * (x - x_list[i]))))

                T = float(T_c - (V ** 2 / (2 * c_p)))

                a = float(sp.sqrt(gamma * R * T))

                M = float(V / a)

                v = float((math.sqrt((gamma + 1) / (gamma - 1))) * sp.atan(
                    math.sqrt((gamma - 1) * (M ** 2 - 1) / (gamma + 1))) - sp.atan(math.sqrt(M ** 2 - 1)))

                K_neg = float(theta + v)
                K_pos = float(theta - v)

                mu = float(sp.asin(1 / M))

                kkk = 0
                # --------------------iterations
                for iteration in range(1, numiter + 1, 1):


                    if iteration == numiter + 1:
                        kkk = kkk + numiter + 1

                    iteration_num = kkk + iteration

                    x_2 = float((y_list[i] - y_list[i + 1] + x_list[i + 1] * 0.5 * (
                            sp.tan(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) + sp.tan(theta - mu)) -
                                 x_list[i] * 0.5 * (sp.tan(theta_xx_rad_list[i] + mu_xx_rad_list[i]) + sp.tan(
                                theta + mu))) / (0.5 * (
                            sp.tan(theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1]) - sp.tan(
                        theta_xx_rad_list[i] + mu_xx_rad_list[i]) + sp.tan(theta - mu) - sp.tan(theta + mu))))

                    y_2 = float(y_list[i] + (x - x_list[i]) * 0.5 * (
                            sp.tan(theta_xx_rad_list[i] + mu_xx_rad_list[i]) + sp.tan(theta + mu)))

                    A_2A = float(
                        (sp.tan(mu_xx_rad_list[i]) * sp.sin(mu_xx_rad_list[i]) * sp.sin(theta_xx_rad_list[i])) / (
                                y_list[i] * sp.cos(theta_xx_rad_list[i] + mu_xx_rad_list[i])))

                    A_2C = float((sp.tan(mu) * sp.sin(mu) * sp.sin(theta)) / (y * sp.cos(theta + mu)))

                    B_2B = float((sp.tan(mu_xx_rad_list[i + 1]) * sp.sin(mu_xx_rad_list[i + 1]) * sp.sin(
                        theta_xx_rad_list[i + 1])) / (y_list[i + 1] * sp.cos(
                        theta_xx_rad_list[i + 1] - mu_xx_rad_list[i + 1])))

                    B_2C = float((sp.tan(mu) * sp.sin(mu) * sp.sin(theta)) / (y * sp.cos(theta - mu)))

                    A_2 = float((2 / (sp.tan(mu_xx_rad_list[i]) + sp.tan(mu))) * (
                            ((2 * V_xx_list[i]) / (V_xx_list[i] + V)) + ((A_2A + A_2C) / 2) * (x - x_list[i])))

                    B_2 = float((2 / (sp.tan(mu_xx_rad_list[i + 1]) + sp.tan(mu))) * (
                            ((2 * V_xx_list[i + 1]) / (V_xx_list[i + 1] + V)) + ((B_2B + B_2C) / 2) * (
                            x - x_list[i + 1])))

                    V_2 = float((A_2 + B_2 + theta_xx_rad_list[i + 1] - theta_xx_rad_list[i]) / (4 * (
                            (V_xx_list[i] + V) ** (-1) * (sp.tan(mu_xx_rad_list[i]) + sp.tan(mu)) ** (-1) + (
                            V_xx_list[i + 1] + V) ** (-1) * (
                                    sp.tan(mu_xx_rad_list[i + 1]) + sp.tan(mu)) ** (-1))))

                    theta_2 = float(theta_xx_rad_list[i] + (2 / (sp.tan(mu_xx_rad_list[i]) + sp.tan(mu))) * (
                            ((2 * (V_2 - V_xx_list[i])) / (V + V_xx_list[i])) - ((A_2A + A_2C) / 2) * (
                            x - x_list[i])))

                    T_2 = float(T_c - (V_2 ** 2 / (2 * c_p)))

                    a_2 = float(sp.sqrt(gamma * R * T_2))

                    M_2 = float(V_2 / a_2)

                    mu_2 = float(sp.asin(1 / M_2))

                    v_2 = float((math.sqrt((gamma + 1) / (gamma - 1))) * sp.atan(
                        math.sqrt((gamma - 1) * (M_2 ** 2 - 1) / (gamma + 1))) - sp.atan(math.sqrt(M_2 ** 2 - 1)))

                    K_neg_2 = float(theta_2 + v_2)
                    K_pos_2 = float(theta_2 - v_2)

                    xdiff = float(x - x_2)
                    if xdiff == 0:
                        xerr = float(0)
                    else:
                        xerr = float(x_2 / xdiff)

                    ydiff = float(y - y_2)
                    if ydiff == 0:
                        yerr = float(0)
                    else:
                        yerr = float(y_2 / ydiff)

                    thetadiff = float(theta - theta_2)
                    if thetadiff == 0:
                        thetaerr = float(0)
                    else:
                        thetaerr = float(theta_2 / thetadiff)

                    Vdiff = float(V - V_2)
                    if Vdiff == 0:
                        Verr = float(0)
                    else:
                        Verr = float(V_2 / Vdiff)

                    x = x_2
                    y = y_2
                    V = V_2
                    theta = theta_2
                    T = T_2
                    a = a_2
                    M = M_2
                    mu = mu_2
                    v = v_2
                    K_neg = K_neg_2
                    K_pos = K_pos_2

            x_list[i] = x
            y_list[i] = y
            V_xx_list[i] = V
            theta_xx_rad_list[i] = theta
            v_PM_xx_rad_list[i] = v
            K_neg_xx_rad_list[i] = K_neg
            K_pos_xx_rad_list[i] = K_pos
            M_xx_list[i] = M
            mu_xx_rad_list[i] = mu
            T_xx_list[i] = T



            plt.plot(x_list, y_list, 'k.', markersize=1)

            n = (z - j + 1) + f


            f = f + z + 1 - i
            sheet.cell(row=n + 1, column=1).value = math.degrees(K_neg_xx_rad_list[i])
            sheet.cell(row=n + 1, column=2).value = math.degrees(K_pos_xx_rad_list[i])
            sheet.cell(row=n + 1, column=3).value = math.degrees(theta_xx_rad_list[i])
            sheet.cell(row=n + 1, column=4).value = math.degrees(v_PM_xx_rad_list[i])
            sheet.cell(row=n + 1, column=5).value = M_xx_list[i]
            sheet.cell(row=n + 1, column=6).value = math.degrees(mu_xx_rad_list[i])
            sheet.cell(row=n + 1, column=7).value = V_xx_list[i]
            sheet.cell(row=n + 1, column=8).value = x_list[i]
            sheet.cell(row=n + 1, column=9).value = y_list[i]
            sheet.cell(row=n + 1, column=11).value = T_xx_list[i]

    counter = 1

    sheet.cell(row=1, column=12).value = "x nozzle wall"
    sheet.cell(row=1, column=13).value = "y nozzle wall"

    d = 0
    for q in range(0, z, 1):


        if q == 0:

            x = float((-x_list[q] * sp.tan(theta_xx_rad_list[q] + mu_xx_rad_list[q]) + y_list[q] - r_t) / (
                    sp.tan(theta_max_rad) - sp.tan(theta_xx_rad_list[q] + mu_xx_rad_list[q])))
            y = float(sp.tan(theta_xx_rad_list[q] + mu_xx_rad_list[q]) * (x - x_list[q]) + y_list[q])
            theta = float((theta_max_rad + theta_xx_rad_list[q]) / 2)

        else:

            x = float((-x_list[q] * sp.tan(theta_xx_rad_list[q] + mu_xx_rad_list[q]) + y_list[q] + x_list[
                q - 1] * sp.tan(theta) - y_list[q - 1]) / (
                              sp.tan(theta) - sp.tan(theta_xx_rad_list[q] + mu_xx_rad_list[q])))
            y = float(sp.tan(theta_xx_rad_list[q] + mu_xx_rad_list[q]) * (x - x_list[q]) + y_list[q])
            theta = float((theta + theta_xx_rad_list[q]) / 2)

        x_list[q] = x
        y_list[q] = y



        d = d + z + 1 - q

        sheet.cell(row=d + 1, column=1).value = math.degrees(K_neg_xx_rad_list[q])
        sheet.cell(row=d + 1, column=2).value = math.degrees(K_pos_xx_rad_list[q])
        sheet.cell(row=d + 1, column=3).value = math.degrees(theta_xx_rad_list[q])
        sheet.cell(row=d + 1, column=4).value = math.degrees(v_PM_xx_rad_list[q])
        sheet.cell(row=d + 1, column=5).value = M_xx_list[q]
        sheet.cell(row=d + 1, column=6).value = math.degrees(mu_xx_rad_list[q])
        sheet.cell(row=d + 1, column=7).value = V_xx_list[q]
        sheet.cell(row=d + 1, column=8).value = x_list[q]
        sheet.cell(row=d + 1, column=9).value = y_list[q]

        sheet.cell(row=counter + 1, column=12).value = x
        sheet.cell(row=counter + 1, column=13).value = y

        counter += 1

    sheet.cell(row=1, column=15).value = "x_conv"
    sheet.cell(row=1, column=16).value = "y_conv"

    x_conv_list = []
    y_conv_list = []

    counter = 1
    for beta in np.arange(-90, -136, -1):


        x_conv = float(1.5 * r_t * sp.cos(math.radians(beta)))
        y_conv = float(1.5 * r_t * sp.sin(math.radians(beta)) + 1.5 * r_t + r_t)

        sheet.cell(row=counter + 1, column=15).value = x_conv
        sheet.cell(row=counter + 1, column=16).value = y_conv

        x_conv_list.append(x_conv)
        y_conv_list.append(y_conv)

        counter += 1

    outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="black")
    outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)

    #outbox.insert(END,"All data from calculation has been exported in 'export1.xlsx'. If another calculation is initiated, the data from the previous calculation will be overwritten." + '\n')
    #outbox.see("end")

    x_line = ((r_c - y_conv) / sp.tan(math.radians(-45))) + x_conv

    x_conv_list.insert(45, x_line)
    y_conv_list.insert(45, r_c)

    plt.plot(x_conv_list, y_conv_list, 'ko', linewidth=2, linestyle='-', markersize=2)

    x_list.insert(0, 0)
    y_list.insert(0, r_t)



    plt.plot(x_list, y_list, 'ko', linewidth=2, linestyle='-', markersize=4)
    plt.ylabel("Radius [m]")
    plt.xlabel("Length [m]")

    plt.axis('equal')

    rerr = float((r_e - y_list[z]) / y_list[z] * 100)

    exhaust_radius.insert(0, y_list[z])
    nozzle_length.insert(0, x_list[z])
    isen_exhaust_radius.insert(0, r_e)
    error.insert(0, rerr)

    textstr = "Combustion Pressure: "+str(p_c)+" [Pa]\n" \
              "Combusiton Temperature: "+str(T_c)+" [K]\n" \
              "Ratio of Spec. Heat: "+str(gamma)+" [-]\n" \
              "Specific Gas Constant: "+str(R)+" [J/kgK]\n" \
              "Throat Radius: "+str(r_t)+" [m]\n" \
              "Combustion Chamber Radius: "+str(r_c)+" [m]\n" \
              "Number of Char. Lines: "+str(num_of_char_lines)+" \n" \
              "Number of Iterations: "+str(numiter)+" \n" \
              "Initial Expansion Angle: "+str(theta_max_deg)+" [deg]\n" \
              "Characteristic Net Length: "+str(L)+" [m]\n" \
              "Pressure Distribution Type: "+str(pressure_type)+" \n" \
              "Cubic q-factor: "+str(q_factor_value)+" [-]\n"

    exhauststr = "Exhaust Mach Number: " + str(M_e) + " [-]\n" \
              "Exhaust velocity: " + str(u_e) + " [m/s]\n" \
              "Exhaust Pressure: " + str(p_e) + " [Pa]\n" \
              "Exhaust Temperature: " + str(T_e) + " [K]\n" \
              "Exhaust Density: " + str(rho_e) + " [kg/m^3]\n" \
              "MOC Exahust Radius: " + str(y_list[z]) + " [m]\n" \
              "Isentropic Exhaust Radius: " + str(r_e) + " [m]\n" \

    T_t = float(T_c*(2/(gamma+1)))
    rho_t = float(rho_c*(2/(gamma+1))**(1/(gamma-1)))

    throatstr = "Throat Mach Number: "+str(M_t)+" [-]\n" \
                "Throat velocity: " + str(float(sp.sqrt(gamma*R*T_t))) + " [m/s]\n" \
                "Throat Pressure: " + str(p_t) + " [Pa]\n" \
                "Throat Temperature: " + str(T_t) + " [K]\n" \
                "Throat Density: " + str(rho_t) + " [kg/m^3]\n" \
                "Initial Expansion Angle: " + str(theta_max_deg) + " [deg]\n" \
                "Isentropic Throat Radius: " + str(r_t) + " [m]\n" \

    combstr = "Combustion Pressure: " + str(p_c) + " [Pa]\n" \
                "Combustion Temperature: " + str(T_c) + " [K]\n" \
                "Combustion Density: " + str(rho_c) + " [kg/m^3]\n" \
                "Combustion Chamber Radius: " + str(r_c) + " [m]\n" \

    lengthstr = "Characteristic Net Length: "+str(L)+" [m] \n" \
                                                     "Nozzle Length: "+str(x_list[z])+" [m]"

    plt.title("AXIMOC Nozzle with "+str(num_of_char_lines)+" char. lines; "+str(numiter)+" iterations. Pressure type: "+str(pressure_type)+"; Expansion angle: "+str(theta_max_deg)+" [deg].")


    plt.text(x_list[z-round(num_of_char_lines/4)], y_list[z], exhauststr, fontsize=8)
    plt.text(x_list[z-round(num_of_char_lines)], y_list[z], throatstr, fontsize=8)
    plt.text(x_line, -0.06, combstr, fontsize=8)
    plt.text(x_list[round(z/3)], -0.025, lengthstr, fontsize=8)

    book.save('export1.xlsx')
    plt.savefig('MOC_Picture1.png')

    outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="black")
    outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)

    outbox.insert(END,"All data from calculation has been exported in 'export1.xlsx' located in the floder of the app. If another calculation is initiated, the data from the previous calculation will be overwritten." + '\n')
    outbox.see("end")

    return


def calculate():

    global outbox

    if all((len(p_e.get()), len(F_T.get()), len(gamma.get()), len(M_e.get()), len(T_c.get()),len(R.get()) ,len(num_of_char_lines.get()), len(numiter.get()), len(theta_max_rad.get()), len(L.get()))) == 0:

        outbox.insert(END, "WARNING: Missing Input! One or more input entries are empty"+'\n')
        outbox.see("end")
        return

    try:
        float(p_e.get()) and float(F_T.get()) and float(gamma.get()) and float(M_e.get()) and float(T_c.get()) and float(R.get()) #and float(r_c.get())
    except ValueError:
        outbox.insert(END, "WARNING: Invalid Input! Input entries in the first column require a positive decimal or integer, not a string." + '\n')
        outbox.see("end")

        return
    try:
        int(num_of_char_lines.get()) and int(numiter.get())

    except ValueError:
        outbox.insert(END, "WARNING: Invalid Input! Number of characteristic lines and number of iterations entries must be positive integers, not strings or floats." + '\n')
        outbox.see("end")

        return
    try:
        float(theta_max_rad.get()) and float(L.get())

    except ValueError:
        outbox.insert(END, "WARNING: Invalid Input! Initial expansion angle and characteristic net length must be positive decimals or integers, not strings." + '\n')
        outbox.see("end")

        return


    if float(p_e.get()) <= 0 or float(F_T.get()) <= 0 or float(gamma.get())<= 0 or float(M_e.get())<= 0 or float(T_c.get())<= 0 or float(R.get())<= 0 or float(num_of_char_lines.get()) <= 0 or float(numiter.get()) <= 0 or float(theta_max_rad.get()) <= 0 or float(L.get())<= 0: # or float(r_c.get()) <= 0
        outbox.insert(END,"WARNING: Negative Input! One or more input entries are negative or zero." + '\n')
        outbox.see("end")
        return


    outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="green4")
    outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)

    def procedure():

        global outbox
        exhaust_radius.delete(0, END)
        nozzle_length.delete(0, END)
        isen_exhaust_radius.delete(0, END)
        error.delete(0, END)

        calculatebutton = Button(tab1, text="Calculate",padx=35, pady=10, command=calculate, state=DISABLED).grid(row=11, column=3, rowspan=2, columnspan=10, sticky=W)

        if r_cset.get() == 1:

            outbox.insert(END, "No recommendations to make at this time." + '\n')
            outbox.see("end")
            outbox.insert(END, "Calculating..." + '\n')
            outbox.see("end")

            AXIMOCdesigner(p_e=float(p_e.get()), F_T=float(F_T.get()), gamma=float(gamma.get()), M_e=float(M_e.get()), T_c=float(T_c.get()), R=float(R.get()), r_c=(r_c.get()), num_of_char_lines=int(num_of_char_lines.get()), numiter=int(numiter.get()), theta_max_rad=float(theta_max_rad.get()), L=float(L.get()), pressure_type='parabolic', q_factor=float(q_factor_value))

            outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="black")
            outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)

            outbox.insert(END,"All data from calculation has been exported in 'export1.xlsx' located in the floder of the app. If another calculation is initiated, the data from the previous calculation will be overwritten." + '\n')
            outbox.see("end")

            calculatebutton = Button(tab1, text="Calculate", padx=35, pady=10, command=calculate).grid(row=11, column=3,rowspan=2,columnspan=10,sticky=W)


        if r_cset.get() == 0:

            outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="red")
            outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)

            if len(r_c.get()) == 0:
                outbox.insert(END, "WARNING: Missing Input! One or more input entries are empty" + '\n')
                outbox.see("end")

                calculatebutton = Button(tab1, text="Calculate", padx=35, pady=10, command=calculate).grid(row=11,column=3,rowspan=2,columnspan=10,sticky=W)

                return

            AXIMOCdesigner(p_e=float(p_e.get()), F_T=float(F_T.get()), gamma=float(gamma.get()), M_e=float(M_e.get()), T_c=float(T_c.get()), R=float(R.get()), r_c=float(r_c.get()), num_of_char_lines=int(num_of_char_lines.get()), numiter=int(numiter.get()), theta_max_rad=float(theta_max_rad.get()), L=float(L.get()), pressure_type='parabolic', q_factor=float(q_factor_value))

        outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="red2")
        outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)
        outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="black")
        outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)

        outbox.insert(END,"All data from calculation has been exported in 'export1.xlsx' located in the folder of the app. If another calculation is initiated, the data from the previous calculation will be overwritten. A plot has also been saved in the same location as a .png file" + '\n')
        outbox.see("end")
        return

    a2 = thrd.Thread(name='procedure', target=procedure)
    global calcnum
    calcnum += 1
    if calcnum == 1:
        a2.start()
    else:
        a2.start()

outbox = Text(tab1, borderwidth=2, width=90, height=10, fg="red2")
outbox.grid(row=26, column=0, columnspan=100, padx=5, pady=40)


p_e = Entry(tab1, borderwidth=1, width = 20)
p_e.grid(row=5, column=1, pady=5)

F_T = Entry(tab1, borderwidth=1, width = 20)
F_T.grid(row=6, column=1, pady=5)

gamma = Entry(tab1, borderwidth=1, width = 20)
gamma.grid(row=7, column=1, pady=5)

M_e = Entry(tab1, borderwidth=1, width = 20)
M_e.grid(row=8, column=1, pady=5)

T_c = Entry(tab1, borderwidth=1, width = 20)
T_c.grid(row=9, column=1, pady=5)

R = Entry(tab1, borderwidth=1, width = 20)
R.grid(row=10, column=1, pady=5)

r_c = Entry(tab1, borderwidth=1, width = 20)
r_c.grid(row=12, column=1, pady=5)

#-----------------------------------------------------------------------------------------------------------------------
num_of_char_lines = Entry(tab1, borderwidth=1, width = 20)
num_of_char_lines.grid(row=5, column=3, pady=5, columnspan=2)

numiter = Entry(tab1, borderwidth=1, width = 20)
numiter.grid(row=6, column=3, pady=5, columnspan=2)

theta_max_rad = Entry(tab1, borderwidth=1, width = 20)
theta_max_rad.grid(row=7, column=3, pady=5, columnspan=2)

L = Entry(tab1, borderwidth=1, width = 20)
L.grid(row=8, column=3, pady=5, columnspan=2)

q_factor = Entry(tab1, borderwidth=1, width = 20, state=DISABLED)
q_factor.grid(row=10, column=3, pady=5, columnspan=2)

Label(tab1, text="Exhaust Pressure [Pa]").grid(row=5, pady=5, column=0, sticky=E)
Label(tab1, text="Thrust [N]").grid(row=6, pady=5, column=0, sticky=E)
Label(tab1, text="Ratio of Specific Heat [-]").grid(row=7, pady=5, column=0, sticky=E)
Label(tab1, text="Exhaust Mach Number [-]").grid(row=8, pady=5, column=0, sticky=E)
Label(tab1, text="Combustion Temperature [K]").grid(row=9, pady=5, column=0, sticky=E)
Label(tab1, text="Ideal Gas Constant [J/kgK]").grid(row=10, pady=5, column=0, sticky=E)
Label(tab1, text="Combustion Chamber Radius [m]").grid(row=11, pady=5, column=0, rowspan=2,sticky=E)
#-----------------------------------------------------------------------------------------------------------------------
Label(tab1, text="Number of Char. Lines [#]").grid(row=5, column=2, sticky=E, pady=5)
Label(tab1, text="Number of Iterations [#]").grid(row=6, column=2, sticky=E, pady=5)
Label(tab1, text="Initial Expansion Angle [rad]").grid(row=7, column=2, sticky=E, pady=5)
Label(tab1, text="Characteristic Net Length [m]").grid(row=8, column=2, sticky=E, pady=5)
Label(tab1, text="Pressure Distribution Type").grid(row=9, column=2, sticky=E, pady=5)
Label(tab1, text="Cubic q-factor [-]").grid(row=10, column=2, sticky=E, pady=5)

parabolic = Button(tab1, text="Parabolic", padx=5, bg="green3", fg="gray99", state=DISABLED, command=parchoice).grid(row=9, column=3, sticky=W) #, command=parchoice
cubic = Button(tab1, text="Cubic", padx=5, bg="gray95", command = cubchoice).grid(row=9, column=3,sticky=E) #, command=cubchoice


progress = Progressbar(tab1, orient=HORIZONTAL, length=700, mode='determinate')
progress.grid(row=20, column=0, columnspan=100, padx=30)
perbox = Entry(tab1, borderwidth=0, width=10, bg="grey95")
perbox.grid(row=19, column=2, pady=5)

remaining = Entry(tab1, borderwidth=1, width=7, bg="grey85", fg="green3")
remaining.grid(row=19, column=0, sticky=E)
Label(tab1, text="Remaining Lines:").grid(row=19)

calculatebutton = Button(tab1, text="Calculate",padx=35, pady=10, command=calculate).grid(row=11, column=3, rowspan=2, columnspan=10, sticky=W)


#-----------------------------------------------------------------------------------------------------------------------

results = Label(tab1, text="_________________________________Results________________________________", font=("Calibri", 15), pady=30).grid(row=22,column=0,columnspan=100)


Label(tab1, text="MOC Exhaust Radius [m]").grid(row=24, column=0, sticky=E, pady=5)
Label(tab1, text="MOC Nozzle Length [m]").grid(row=24, column=2, sticky=E, pady=5)
Label(tab1, text="Isentropic Exhaust Radius [m]").grid(row=25, column=0, sticky=E, pady=5)
Label(tab1, text="Exhaust Radius Error [%]").grid(row=25, column=2, sticky=E, pady=5)

exhaust_radius = Entry(tab1, borderwidth=1, width = 20, bg="grey85", fg="green3")
exhaust_radius.grid(row=24, column=1, sticky=E, pady=5)

nozzle_length = Entry(tab1, borderwidth=1, width = 20, bg="grey85", fg="green3")
nozzle_length.grid(row=24, column=3, sticky=E, pady=5)

isen_exhaust_radius = Entry(tab1, borderwidth=1, width = 20, bg="grey85", fg="green3")
isen_exhaust_radius.grid(row=25, column=1, sticky=E, pady=5)

error = Entry(tab1, borderwidth=1, width = 20, bg="grey85", fg="green3")
error.grid(row=25, column=3, sticky=E, pady=5)

threetimes = Checkbutton(tab1, text="R_c = 1.73*R_t", command=r_cchoice, variable=r_cset).grid(row=11, pady=5, column=1, sticky=W)

Label(tab1, text="Dialogue Box", bg="white").grid(row=26, column=0, sticky=N, pady=21)

#help-------------------------------------------------------------------------------------------------------------------

Label(tab2, text="Help, Instructions, and Troubleshooting", font=("Calibri", 20)).grid(row=0,column=0, pady=30, columnspan=2)

Label(tab2, text="How to Use: ", font=("Calibri", 14), justify=LEFT).grid(row=1,column=1,columnspan=100, pady=10, sticky=W)

Label(tab2, text="In order to initiate a calculation, all input entries must be filled with decimals or integers. There \n"
                 "are some special entries such as the combustion chamber radius entry and the cubic q-factor entry.\n"
                 "Since isentropic theory assumes a zero-flow-velocity at the stagnation point (combustion chamber) a\n"
                 " geometry cannot be calcualted using isentropic theory. Checking the checkbox initiates an approximate\n"
                 " combustion radius which equats to the product of the throat radius and the square root of 3.\n "
                 " If the checkbox is unchecked, a combustion radius must be manually input.In the second column of\n "
                 "inputs, the AXIMOC requirements must be entered. The number of char. lines and iterations dictate the \n"
                 "accuracy of the result and the performance of the programme. The initial expansion angle dictates the angle\n"
                 " at which the first wall segment (from the throat point to the next wall segment) is relative to the x-axis. \n"
                 "The pressure distribution type is by default set to parabolic which is considered a relatively accurate\n"
                 " analytical pressure model. The cubic pressure distribution requires an additional input of a q-factor\n"
                 " which is in the range of -5 to +5. Any inputs beyond that range would produce unrealistic pressure \n"
                 "distributions. Once all entries have been filled, the calculation can be initiated by pressing the \n"
                 "'Calculate' button. The 'Remaining Lines' field will show the remaining characteristic lines and the \n"
                 "progress bar will indicate the progress of the calcualtion. When complete, a new window will be \n"
                 "automatically opened in which the nozzle graph and calculation results can be found. The photo in that\n"
                 " new window can be saved. Furthermore, data from the numerical computation is exported into an excel file\n"
                 " named 'export1.xlsx' which can be found in the main folder of the programme.In the results section of the\n"
                 " main window 4 results are displayed for conveniance. If needed, the input entries can then be modified and a \n"
                 "new calculation can be initialized (Note: To initialize a new calculation the graph window must be closed!). \n"
                 "The Dialogue Box at the bottom will print warnings and cautions when needed (if a mistake in entries is made).\n", font=("Calibri", 11), justify=LEFT).grid(row=2,column=1)

Label(tab2, text="Possible Problems: ", font=("Calibri", 14), justify=LEFT).grid(row=3,column=1,columnspan=100, pady=10, sticky=W)

Label(tab2, text="There are known issues which cause the software to crash. This will most commonly occur when more than 3 \n"
                 "calculations have been initialized sequentially.If this happens, the prgoramme must be closed and reopened. \n"
                 "This issue will be fixed in a later version.In cases where the 'Calculate' button is pressed and nothing happens,\n"
                 " an error must have occured. This would be caused by an invalid input which was not caught by the software.\n"
                 "Checking the command prompt of the programme would confirm such a case. To fix such a conditions, fix the \n"
                 "wrong input or restart the programme.\n", font=("Calibri", 11), justify=LEFT).grid(row=4,column=1)

Label(tab2, text="Links: ", font=("Calibri", 14), justify=LEFT).grid(row=5,column=1,columnspan=100, pady=10, sticky=W)

Label(tab2, text="This programme was developed as a part of a paper on Axisymmetric Method of Characteristics. \n"
                 "For more information on the methodology and procedures of the calculator review 'NAME OF PAPER' by \n"
                 "Kyril Palaveev and Dr. Torsten Schenkel.\n", font=("Calibri", 11), justify=LEFT).grid(row=6,column=1)


tab1.update()
tab1.mainloop()



